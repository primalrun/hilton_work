import redshift_connector as rc
import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import filedialog
import tkinter as tk
import os
import pandas as pd
import datetime
from dateutil.relativedelta import relativedelta


def connect_rs(host_p, database_p, user_p, password_p):
    conn = rc.connect(
        host=host_p
        , database=database_p
        , user=user_p
        , password=password_p
    )
    conn.autocommit = True
    return conn


def create_excel_workbook(file_name_in):
    wb = openpyxl.Workbook()
    wb.save(file_name_in)
    wb.close()


def clean_excel_file(wb_in, file_excel_p):
    for sheet in wb_in.sheetnames:
        if str(sheet)[0:5] == 'Sheet':
            wb_in.remove(wb_in[sheet])

    for sheet in wb_in.sheetnames:
        ws = wb_in[sheet]
        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max([len(str(cell.value) or "") for cell in column_cells])
            max_length = max_length + 3
            ws.column_dimensions[column_letter].width = max_length
            ws.freeze_panes = 'A2'
    wb_in.save(file_excel_p)


def load_excel_file_and_clean(wb_file_name_p):
    wb = openpyxl.load_workbook(wb_file_name_p)
    clean_excel_file(wb)
    wb.save(wb_file_name_p)
    wb.close()


def select_input_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()

    if file_path:
        return file_path
    else:
        return None


def write_df_to_excel(df_p, file_name_p, sheet_name_p):
    if os.path.exists(file_name_p):
        os.remove(file_name_p)
    df_p.to_excel(file_name_p, index=False, sheet_name=sheet_name_p)
    os.startfile(file_name_p)


def write_df_to_csv(df_p, file_name_p):
    if os.path.exists(file_name_p):
        os.remove(file_name_p)
    df_p.to_csv(file_name_p, index=False)
    os.startfile(file_name_p)


def delete_file_if_exists(file_path_p):
    if os.path.exists(file_path_p):
        os.remove(file_path_p)


def convert_period_yyyymm_int_to_start_date_and_end_date(period_yyyymm_int_p):
    date_start = datetime.datetime.strptime(str(period_yyyymm_int_p), '%Y%m').date()
    date_end = date_start + relativedelta(months=1) - relativedelta(days=1)
    date_start_str = date_start.strftime('%Y-%m-%d')
    date_end_str = date_end.strftime('%Y-%m-%d')
    date_start_yyyymmdd = date_start.strftime('%Y%m%d')
    date_end_yyyymmdd = date_end.strftime('%Y%m%d')
    return date_start, date_end, date_start_str, date_end_str, date_start_yyyymmdd, date_end_yyyymmdd


def replace_sql_from_dict(sql_p, dict_replacement_p):
    sql = sql_p
    for key, value in dict_replacement_p.items():
        sql = sql.replace(key, value)
    return sql


def convert_df_columns_to_numeric(df_in, columns_p, number_of_decimals_p):
    dict_decimal = {}
    for i in range(len(columns_p)):
        dict_decimal[columns_p[i]] = number_of_decimals_p[i]
    df_in[columns_p] = df_in[columns_p].apply(pd.to_numeric, errors='coerce')
    for key, value in dict_decimal.items():
        df_in[key] = df_in[key].round(value)
    return df_in.copy()


def clean_df(df_in):
    if 'country_cd' in df_in.columns:
        df_in['country_cd'] = df_in['country_cd'].fillna('Unknown')
    if 'op_area_level2_desc' in df_in.columns:
        df_in['op_area_level2_desc'] = df_in['op_area_level2_desc'].fillna('Unknown')
    df_out = df_in.copy()
    return df_out


def append_df_to_existing_excel_workbook(df_p, workbook_path_p, sheet_name_p):
    with pd.ExcelWriter(workbook_path_p, mode='a', engine='openpyxl') as writer:
        df_p.to_excel(writer, sheet_name=sheet_name_p, index=False)


def calc_variance_pct(var_amt, base_amt):
    if base_amt == 0:
        if var_amt == 0:
            return 0
        if var_amt > 0:
            return 1
        if var_amt < 0:
            return -1
    if base_amt < 0:
        return - var_amt / base_amt
    if base_amt > 0:
        return var_amt / base_amt


def add_variance_columns(df_in, measure_columns_p, charge_category_p, version_compare_from, version_compare_to):
    df_in[measure_columns_p] = df_in[measure_columns_p].fillna(0)

    df_in['room nights var amt'] = df_in[f"room nights {version_compare_to}"] - df_in[f"room nights {version_compare_from}"]
    df_in[f"{charge_category_p} rev var amt"] = df_in[f"{charge_category_p} rev {version_compare_to}"] - df_in[f"{charge_category_p} rev {version_compare_from}"]

    df_in['room nights var pct'] = df_in.apply(
        lambda x: calc_variance_pct(x['room nights var amt'], x[f"room nights {version_compare_from}"]), axis=1)
    df_in[f"{charge_category_p} rev var pct"] = df_in.apply(
        lambda x: calc_variance_pct(x[f"{charge_category_p} rev var amt"], x[f"{charge_category_p} rev {version_compare_from}"]), axis=1)
    return df_in.copy()


def calc_variance_pct_between_from_and_to(from_amt, to_amt):
    var_amt = to_amt - from_amt
    if from_amt == 0:
        if var_amt == 0:
            return 0
        if var_amt > 0:
            return 1
        if var_amt < 0:
            return -1
    if from_amt < 0:
        return - var_amt / from_amt
    if from_amt > 0:
        return var_amt / from_amt



def sort_df(df_p, sort_columns_p):
    column_order = []
    for col in sort_columns_p:
        if col in df_p.columns:
            column_order.append(col)
    df_p = df_p[column_order]
    return df_p.copy()


def format_excel_data(wb_p, sheet_format_dict_p, file_excel_p):
    for sheet in sheet_format_dict_p.keys():
        ws = wb_p[sheet]
        format_dict = sheet_format_dict_p[sheet][0]
        row_start, row_end = sheet_format_dict_p[sheet][1:]
        if row_start is None:
            row_start = 2
        if row_end is None:
            row_end = ws.max_row
        for fmt in format_dict:
            for c in format_dict[fmt]:
                for r in range(row_start, row_end + 1):
                    ws.cell(row=r, column=c).number_format = fmt
    wb_p.save(file_excel_p)


def rename_df_columns(df_p, columns_p, version_compare_to_p, version_compare_from_p, charge_category_p):
    rename_dict = {
        f"room_nights {version_compare_to_p}": f"room nights {version_compare_to_p}"
        ,f"room_nights {version_compare_from_p}": f"room nights {version_compare_from_p}"
        ,f"chkout_{charge_category_p}_usd_amt {version_compare_to_p}": f"{charge_category_p} rev {version_compare_to_p}"
        ,f"chkout_{charge_category_p}_usd_amt {version_compare_from_p}": f"{charge_category_p} rev {version_compare_from_p}"
    }
    df_p = df_p.rename(columns=rename_dict)
    return df_p.copy()


def sort_excel_sheets_from_list(wb_p, sheet_order_p, file_excel_p):
    for sheet_name in reversed(sheet_order_p): # move sheets to front of workbook from last to first
        ws = wb_p[sheet_name]
        wb_p.move_sheet(ws, 0 - wb_p.sheetnames.index(sheet_name))
    wb_p.save(file_excel_p)

def rename_df_columns_to_uppercase_source(df_p, charge_category_p):
    columns_to_rename = [col for col in list(df_p.columns.values) if charge_category_p in col]
    column_rename = []
    for col in columns_to_rename:
        col_1 = str(col).replace('_', ' ')
        rev_category, rev, source, version = col_1.split(' ')
        source = source.upper()
        col_new = ' '.join([rev_category, rev, source, version])
        column_rename.append(col_new)
    dict_col_replace = dict(zip(columns_to_rename, column_rename))
    df_p = df_p.rename(columns=dict_col_replace)
    return df_p.copy()


def replace_df_column_name_prop_specific(df_p, column_to_exclude, search_char_p, replace_char_p, charge_category_p):
    columns_to_rename = [col for col in list(df_p.columns.values) if col not in column_to_exclude]
    column_rename = []
    for col in columns_to_rename:
        col_1 = str(col).replace('_', ' ')
        if charge_category_p in col:
            rev_category, rev, source, version = col_1.split(' ')
            source = source.upper()
            col_new = ' '.join([rev_category, rev, source, version])
            column_rename.append(col_new)
        else:
            column_rename.append(col_1)
    dict_col_replace = dict(zip(columns_to_rename, column_rename))
    df_p = df_p.rename(columns=dict_col_replace)
    return df_p.copy()


def get_variance_columns_from_measure_columns(column_list_p, charge_category_p):
    measure_abbrev_columns = []
    for col in column_list_p:
        second_space_index = col.index(' ', len(f"{charge_category_p} rev"))
        col_new = col[second_space_index + 1:]
        measure_abbrev_columns.append(col_new)

    variance_columns = []
    dict_variance_combo = {}
    for col in measure_abbrev_columns:
        other_columns = [column for column in measure_abbrev_columns if column != col]
        for other_col in other_columns:
            if f"{col}^{other_col}" not in dict_variance_combo and f"{other_col}^{col}" not in dict_variance_combo:
                var_amt = f"{col} vs. {other_col} var amt"
                var_pct = f"{col} vs. {other_col} var pct"
                variance_columns.append(var_amt)
                variance_columns.append(var_pct)
                dict_variance_combo[f"{col}^{other_col}"] = 1
                dict_variance_combo[f"{other_col}^{col}"] = 1
    return variance_columns



def get_rev_var_pct_distribution(var_pct_col):
    if var_pct_col < -1:
        return '< -100%'
    if var_pct_col == -1:
        return 'equals -100%'
    if var_pct_col > -1 and var_pct_col <= -.9:
        return 'between -100% and -90%'
    if var_pct_col > -.9 and var_pct_col <= -.8:
        return 'between -90% and -80%'
    if var_pct_col > -.8 and var_pct_col <= -.7:
        return 'between -80% and -70%'
    if var_pct_col > -.7 and var_pct_col <= -.6:
        return 'between -70% and -60%'
    if var_pct_col > -.6 and var_pct_col <= -.5:
        return 'between -60% and -50%'
    if var_pct_col > -.5 and var_pct_col <= -.4:
        return 'between -50% and -40%'
    if var_pct_col > -.4 and var_pct_col <= -.3:
        return 'between -40% and -30%'
    if var_pct_col > -.3 and var_pct_col <= -.2:
        return 'between -30% and -20%'
    if var_pct_col > -.2 and var_pct_col <= -.1:
        return 'between -20% and -10%'
    if var_pct_col > -.1 and var_pct_col <= -.08:
        return 'between -10% and -8%'
    if var_pct_col > -.08 and var_pct_col <= -.06:
        return 'between -8% and -6%'
    if var_pct_col > -.06 and var_pct_col <= -.04:
        return 'between -6% and -4%'
    if var_pct_col > -.04 and var_pct_col <= -.02:
        return 'between -4% and -2%'
    if var_pct_col > -.02 and var_pct_col <= 0:
        return 'between -2% and 0%'
    if var_pct_col == 0:
        return 'equals 0%'
    if var_pct_col > 0 and var_pct_col <= .02:
        return 'between 0% and 2%'
    if var_pct_col > .02 and var_pct_col <= .04:
        return 'between 2% and 4%'
    if var_pct_col > .04 and var_pct_col <= .06:
        return 'between 4% and 6%'
    if var_pct_col > .06 and var_pct_col <= .08:
        return 'between 6% and 8%'
    if var_pct_col > .08 and var_pct_col <= .1:
        return 'between 8% and 10%'
    if var_pct_col > .1 and var_pct_col <= .2:
        return 'between 10% and 20%'
    if var_pct_col > .2 and var_pct_col <= .3:
        return 'between 20% and 30%'
    if var_pct_col > .3 and var_pct_col <= .4:
        return 'between 30% and 40%'
    if var_pct_col > .4 and var_pct_col <= .5:
        return 'between 40% and 50%'
    if var_pct_col > .5 and var_pct_col <= .6:
        return 'between 50% and 60%'
    if var_pct_col > .6 and var_pct_col <= .7:
        return 'between 60% and 70%'
    if var_pct_col > .7 and var_pct_col <= .8:
        return 'between 70% and 80%'
    if var_pct_col > .8 and var_pct_col <= .9:
        return 'between 80% and 90%'
    if var_pct_col > .9 and var_pct_col < 1:
        return 'between 90% and 100%'
    if var_pct_col == 1:
        return 'equals 100%'
    if var_pct_col > 1:
        return '> 100%'

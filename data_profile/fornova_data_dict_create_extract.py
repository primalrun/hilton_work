import config
import common_functions as cfx
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import sys
import os

file_version_data_dict_fornova = config.file_version_data_dict_fornova
file_data_dict_fornova = config.file_data_dict_fornova
data_type_normalization_dict = config.data_type_normalization_dict
file_fornova_data_type_overwrite = config.file_fornova_data_type_overwrite
file_data_dict_fornova_processed = config.file_data_dict_fornova_processed

sheet_name_fornova_data_dict = {
    '20250417_1': [
        'Parity Scan Dictionary v2'
        ,'Compliance Dictionary v2'
        ,'Extranet Scan Dictionary v2'
        ,'Test Res Dictionary v2'
        ,'User Data Dictionary'
    ]
}

column_name_fornova_data_dict = {
    '20250417_1': [
        'Item No.'
        ,'Filename'
        ,'Field Name'
        ,'Data Type'
        ,'Max Length (if String)'
        ,'Mandatory (Y/N) for Business Reporting'
        ,'Example'
        ,'Description'
    ]
}

field_data_type_fornova_data_dict = {
    '20250417_1': [
        'int8'
        ,'string'
        ,'string'
        ,'string'
        ,'string'
        ,'string'
        ,'string'
        ,'string'
        ,'string'
        ,'string'
    ]
}


def get_column_header_row(ws, search_cell_range_str, search_str):
    cell_range = ws[search_cell_range_str]
    for row in cell_range:
        for cell in row:
            if cell.value == search_str:
                return cell.row, cell.column


def get_last_data_row(ws, col_letter):
    last_row = 400
    while last_row > 0:
        if ws[f"{col_letter}{last_row}"].value is not None:
            return last_row
        last_row -= 1
    return 0


def get_char_max_length(max_length, data_type):
    if data_type == 'varchar':
        if cfx.is_number(max_length) is True:
            return max_length
        else:
            return None
    else:
        return None

def get_is_required_for_reporting(mandatory_str):
    if mandatory_str.upper() == 'Y':
        return 'yes'
    if mandatory_str.upper() == 'N':
        return 'no'
    if mandatory_str == 'Y (only scans that have a reshop)':
        return 'yes partial, if scans have a reshop'
    if mandatory_str == 'Y (except for clickbait lines - Null)':
        return 'yes partial, if clickbait lines are not null'
    else:
        return None


sheet_names = sheet_name_fornova_data_dict[file_version_data_dict_fornova]
column_names = column_name_fornova_data_dict[file_version_data_dict_fornova]
data_types = field_data_type_fornova_data_dict[file_version_data_dict_fornova]
column_data_type_dict = dict(zip(column_names, data_types))

df_data_type_overwrite = pd.read_excel(io=file_fornova_data_type_overwrite)
# print(df_data_type_overwrite.head())

wb = openpyxl.load_workbook(filename=file_data_dict_fornova, read_only=True, data_only=True)
df_list = []

for sheet in sheet_names:
    ws = wb[sheet]
    row_nbr_start, col_nbr_start = get_column_header_row(ws, 'A1:A3', 'Item No.')
    row_last = get_last_data_row(ws, 'A')
    header_column_count = len(column_name_fornova_data_dict[file_version_data_dict_fornova])
    row_nbr_end = row_last
    col_nbr_end = (col_nbr_start + header_column_count) - 1
    header_range = ws.iter_rows(min_row=row_nbr_start, max_row=row_nbr_start, min_col=col_nbr_start, max_col=col_nbr_end, values_only=True)
    header_values = [value for row in header_range for value in row]

    #check if column headers match expected header values
    if column_names == header_values:
        print(f'column headers match expected values for sheet {sheet}')
    else:
        print(f'column headers do not match expected values for sheet {sheet}')
        print('process cancelled')
        sys.exit()

    df_iter = cfx.convert_excel_range_to_df(ws, row_nbr_start, row_nbr_end, col_nbr_start, col_nbr_end)
    df_iter = df_iter.astype(dtype=column_data_type_dict)

    # test to write df to excel
    # cfx.write_df_to_excel_temp(df_iter, r'c:\temp\1.xlsx', '1')
    # os.startfile(r'c:\temp\1.xlsx')

    df_list.append(df_iter)

df_data_dict = pd.concat(df_list)
df_data_dict['data_type_normalized'] = df_data_dict.apply(lambda x: cfx.data_type_normalize(x['Data Type'], data_type_normalization_dict), axis=1)
df_data_dict = df_data_dict.merge(df_data_type_overwrite[['Filename', 'Field Name', 'data_type_overwrite']], how='left', on=['Filename', 'Field Name'])
df_data_dict['data_type_new'] = df_data_dict['data_type_overwrite'].combine_first(df_data_dict['data_type_normalized'])
df_data_dict['char_max_length'] = df_data_dict.apply(lambda x: get_char_max_length(x['Max Length (if String)'], x['data_type_new']), axis=1)
df_data_dict['char_max_length'] = df_data_dict['char_max_length'].astype('Int64')
df_data_dict['is_required_for_reporting'] = df_data_dict.apply(lambda x: get_is_required_for_reporting(x['Mandatory (Y/N) for Business Reporting']), axis=1)

# test to write df to excel
# cfx.write_df_to_excel_temp(df_data_dict, r'c:\temp\1.xlsx', '1')
# os.startfile(r'c:\temp\1.xlsx')

cfx.delete_file_if_exists(file_data_dict_fornova_processed)
cfx.write_df_to_excel(df_data_dict, file_data_dict_fornova_processed, 'fornova data dict')
cfx.load_excel_file_and_clean(file_data_dict_fornova_processed)

# add wrap text to columns G, H
wb = openpyxl.load_workbook(filename=file_data_dict_fornova_processed)
ws = wb['fornova data dict']
row_last = ws.max_row
col_last = ws.max_column
wrap_text_range = ws.iter_rows(min_row=2, max_row=row_last, min_col=7, max_col=8)
ws.column_dimensions['G'].width = 40
ws.column_dimensions['H'].width = 40

for row in wrap_text_range:
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

wb.save(file_data_dict_fornova_processed)
wb.close()


print('success')

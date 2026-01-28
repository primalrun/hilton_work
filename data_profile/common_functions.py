import redshift_connector as rc
import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import filedialog
import tkinter as tk
import os
import pandas as pd
import datetime
from dateutil.relativedelta import relativedelta
import polars as pl
import re

def connect_rs(host_p, database_p, user_p, password_p):
    conn = rc.connect(
        host=host_p
        , database=database_p
        , user=user_p
        , password=password_p
    )
    conn.autocommit = True
    return conn


def create_excel_workbook(file_name_in):
    wb = openpyxl.Workbook()
    wb.save(file_name_in)
    wb.close()


def clean_excel_file(wb_in, file_excel_p):
    for sheet in wb_in.sheetnames:
        if str(sheet)[0:5] == 'Sheet':
            wb_in.remove(wb_in[sheet])

    for sheet in wb_in.sheetnames:
        ws = wb_in[sheet]
        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max([len(str(cell.value) or "") for cell in column_cells])
            max_length = max_length + 2
            ws.column_dimensions[column_letter].width = max_length
            ws.freeze_panes = 'A2'
    wb_in.save(file_excel_p)


def load_excel_file_and_clean(wb_file_name_p):
    wb = openpyxl.load_workbook(wb_file_name_p)
    clean_excel_file(wb, wb_file_name_p)
    wb.save(wb_file_name_p)
    wb.close()


def select_input_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()

    if file_path:
        return file_path
    else:
        return None

def write_df_to_excel(df_p, file_name_p, sheet_name_p):
    if os.path.exists(file_name_p):
        os.remove(file_name_p)
    df_p.to_excel(file_name_p, index=False, sheet_name=sheet_name_p)


def write_df_to_excel_temp(df_p, file_name_p, sheet_name_p):
    if os.path.exists(file_name_p):
        os.remove(file_name_p)
    df_p.to_excel(file_name_p, index=False, sheet_name=sheet_name_p)
    os.startfile(file_name_p)


def write_df_to_csv(df_p, file_name_p):
    if os.path.exists(file_name_p):
        os.remove(file_name_p)
    df_p.to_csv(file_name_p, index=False)
    os.startfile(file_name_p)


def delete_file_if_exists(file_path_p):
    if os.path.exists(file_path_p):
        os.remove(file_path_p)


def convert_period_yyyymm_int_to_start_date_and_end_date(period_yyyymm_int_p):
    date_start = datetime.datetime.strptime(str(period_yyyymm_int_p), '%Y%m').date()
    date_end = date_start + relativedelta(months=1) - relativedelta(days=1)
    date_start_str = date_start.strftime('%Y-%m-%d')
    date_end_str = date_end.strftime('%Y-%m-%d')
    date_start_yyyymmdd = date_start.strftime('%Y%m%d')
    date_end_yyyymmdd = date_end.strftime('%Y%m%d')
    return date_start, date_end, date_start_str, date_end_str, date_start_yyyymmdd, date_end_yyyymmdd


def replace_sql_from_dict(sql_p, dict_replacement_p):
    sql = sql_p
    for key, value in dict_replacement_p.items():
        sql = sql.replace(key, value)
    return sql


def convert_df_columns_to_numeric(df_in, columns_p, number_of_decimals_p):
    dict_decimal = {}
    for i in range(len(columns_p)):
        dict_decimal[columns_p[i]] = number_of_decimals_p[i]
    df_in[columns_p] = df_in[columns_p].apply(pd.to_numeric, errors='coerce')
    for key, value in dict_decimal.items():
        df_in[key] = df_in[key].round(value)
    return df_in.copy()



def append_df_to_existing_excel_workbook(df_p, workbook_path_p, sheet_name_p):
    with pd.ExcelWriter(workbook_path_p, mode='a', engine='openpyxl') as writer:
        df_p.to_excel(writer, sheet_name=sheet_name_p, index=False)


def calc_variance_pct(var_amt, base_amt):
    if base_amt == 0:
        if var_amt == 0:
            return 0
        if var_amt > 0:
            return 1
        if var_amt < 0:
            return -1
    if base_amt < 0:
        return - var_amt / base_amt
    if base_amt > 0:
        return var_amt / base_amt


def sort_df(df_p, sort_columns_p):
    column_order = []
    for col in sort_columns_p:
        if col in df_p.columns:
            column_order.append(col)
    df_p = df_p[column_order]
    return df_p.copy()


def format_excel_data(wb_p, sheet_format_dict_p, file_excel_p):
    for sheet in sheet_format_dict_p.keys():
        ws = wb_p[sheet]
        format_dict = sheet_format_dict_p[sheet][0]
        row_start, row_end = sheet_format_dict_p[sheet][1:]
        if row_start is None:
            row_start = 2
        if row_end is None:
            row_end = ws.max_row
        for fmt in format_dict:
            for c in format_dict[fmt]:
                for r in range(row_start, row_end + 1):
                    ws.cell(row=r, column=c).number_format = fmt
    wb_p.save(file_excel_p)



def sort_excel_sheets_from_list(wb_p, sheet_order_p, file_excel_p):
    for sheet_name in reversed(sheet_order_p): # move sheets to front of workbook from last to first
        ws = wb_p[sheet_name]
        wb_p.move_sheet(ws, 0 - wb_p.sheetnames.index(sheet_name))
    wb_p.save(file_excel_p)

def rename_df_columns_to_uppercase_source(df_p, charge_category_p):
    columns_to_rename = [col for col in list(df_p.columns.values) if charge_category_p in col]
    column_rename = []
    for col in columns_to_rename:
        col_1 = str(col).replace('_', ' ')
        rev_category, rev, source, version = col_1.split(' ')
        source = source.upper()
        col_new = ' '.join([rev_category, rev, source, version])
        column_rename.append(col_new)
    dict_col_replace = dict(zip(columns_to_rename, column_rename))
    df_p = df_p.rename(columns=dict_col_replace)
    return df_p.copy()


def convert_excel_range_to_df(ws, row_nbr_start, row_nbr_end, col_nbr_start, col_nbr_end):
    data_range = ws.iter_rows(
        min_row=row_nbr_start
        , max_row=row_nbr_end
        , min_col=col_nbr_start
        , max_col=col_nbr_end
        , values_only=True)
    data_rows = [row for row in data_range]
    headers = data_rows[0]
    data = data_rows[1:]
    df = pd.DataFrame(data, columns=headers)
    return df

def data_type_normalize(data_type, data_type_normalization_dict):
    if data_type.lower() in data_type_normalization_dict:
        return data_type_normalization_dict[data_type.lower()]
    else:
        return data_type.lower()

def is_number(str):
    try:
        float(str)
        return True
    except ValueError:
        return False


def write_pl_df_to_excel_temp(df, file_name):
    if os.path.exists(file_name):
        os.remove(file_name)
    try:
        df.write_excel(file_name)
        os.startfile(file_name)
    except Exception as e:
        print(f"An error occurred: {e}")


def write_pl_df_to_excel(df, file_name, sheet_name):
    try:
        df.write_excel(workbook=file_name, worksheet=sheet_name)
    except Exception as e:
        print(f"An error occurred: {e}")


def pl_count_blank_strings(df: pl.DataFrame) -> pl.DataFrame:
    expressions = []
    for col_name, dtype in df.schema.items():
        if dtype == pl.Utf8 or dtype == pl.String:
            # filter for empty strings, then count
            count_expr = (pl.col(col_name).str.len_chars() == 0).sum().alias(col_name)
            expressions.append(count_expr)
        else:
            # for non-string columns, return 0
            null_expr = pl.lit(None, dtype=pl.UInt32).alias(col_name)
            expressions.append(null_expr)
    return df.select(expressions)


def data_profile_file_in_get_attributes(file_name):
    if '_' in file_name:
        file_split = file_name.split('_')
        if 'correction' in file_name:
            file_type = 'correction'
        else:
            file_type = 'daily'
        try:
            # check if date is integer, possibly in yyyymmdd format
            int(file_split[1])
            # subject, yyyymmdd, file_type
            return file_split[0], file_split[1], file_type
        except ValueError:
            # date is not an integer, try to convert based on known date pattern "Full Month Name-day (with ordinal suffixes)-full year"
            date_str = file_split[1]
            date_without_oridinal_suffix = re.sub(r"(\d+)(st|nd|rd|th)", r'\1', date_str)
            date_object = datetime.datetime.strptime(date_without_oridinal_suffix, '%B-%d-%Y')
            date_yyyymmdd = date_object.strftime('%Y%m%d')
            return file_split[0], date_yyyymmdd, file_type
    else:
        return None


def pl_get_min_length_of_characters(df: pl.DataFrame) -> pl.DataFrame:
    expressions = []
    for col_name, dtype in df.schema.items():
        if dtype == pl.Utf8 or dtype == pl.String:
            # filter for empty strings, then count
            min_char_length_expr = (pl.col(col_name).drop_nulls().str.len_chars()).min().alias(col_name)
            expressions.append(min_char_length_expr)
        else:
            # for non-string columns, return 0
            null_expr = pl.lit(None, dtype=pl.UInt32).alias(col_name)
            expressions.append(null_expr)
    return df.select(expressions)


def get_shortest_string_per_column(df: pl.DataFrame) -> dict:
    """
    Gets the value with the least amount of characters for each string column.
    """
    shortest_values = {}

    # Iterate over string columns
    for col_name in df.columns:
        # Check if the column is a string type
        if df[col_name].dtype == pl.Utf8 or df[col_name].dtype == pl.String:
            # Create a new column with string lengths
            min_len_idx = df[col_name].drop_nulls().str.len_chars().arg_min()

            # Retrieve the value at that index
            # Use pl.element() or select/item() to get the specific value
            shortest_value = df.item(min_len_idx, col_name)

            shortest_values[col_name] = shortest_value
        else:
            shortest_values[col_name] = None
    return shortest_values


def pl_convert_df_to_string(df: pl.DataFrame) -> pl.DataFrame:
    return df.with_columns(
        pl.all().cast(pl.Utf8)
    )

def convert_xlsx_to_csv(excel_file_path, csv_file_path, sheet_name=0):
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
    df.to_csv(csv_file_path, index=False, encoding='utf-8')
    return print(f"Successfully converted {excel_file_path} to {csv_file_path}")



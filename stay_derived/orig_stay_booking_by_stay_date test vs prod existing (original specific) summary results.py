"""
orig_stay_booking_by_stay_date test vs prod existing (original specific) summary results.py

Reads all Excel files matching the pattern:
    orig_stay_booking_by_stay_date test environment vs PROD existing (original specific) (*stay_id *).xlsx

Each file is expected to have a "variance" sheet with paired columns:
    <column_name>_test  and  <column_name>_prod

Produces a summary Excel workbook with:

  Sheet 1 - "Variance Summary":
  - One row per source file
  - One column per unique column that had a variance (across all files)
  - Cell value = number of rows with a variance in that column for that file
  - A "Total Variances" column summing across all columns for the file
  - Zero-variance files are still listed (all counts will be 0)

  Sheet per variance column - "<abbreviated column name> detail":
  - One row per source file where that column had a variance
  - Columns: File, Stay ID, <col>_test, <col>_prod
  - Sheet names are abbreviated to fit Excel's 31-character limit

Usage:
    script.py [directory] [output_file]
"""

import sys
import os
import glob
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── config ──────────────────────────────────────────────────────────────────
VARIANCE_SHEET   = "variance"
FILE_PATTERN     = "orig_stay_booking_by_stay_date test environment vs PROD existing (original specific) (*stay_id *).xlsx"
STAY_ID_RE       = re.compile(r"stay_id (\d+)", re.IGNORECASE)
SHEET_MAX_LEN    = 31
DETAIL_SUFFIX    = " detail"   # appended to every detail sheet name


def extract_stay_id(filename: str) -> str:
    m = STAY_ID_RE.search(filename)
    return m.group(1) if m else os.path.basename(filename)


def make_sheet_name(col_name: str, used: set[str]) -> str:
    """
    Produce a unique Excel sheet name <= 31 chars for '<col_name> detail'.
    Strategy: truncate the column name so the full label fits, then append
    a numeric suffix if the name is already taken.
    Forbidden chars [ ] : * ? / \\ are stripped.
    """
    forbidden = re.compile(r"[\[\]:*?/\\]")
    base = forbidden.sub("", col_name)
    suffix = DETAIL_SUFFIX
    max_base = SHEET_MAX_LEN - len(suffix)
    candidate = (base[:max_base] + suffix).strip()

    if candidate not in used:
        used.add(candidate)
        return candidate

    # Append incrementing number until unique
    for n in range(2, 1000):
        tag = f" {n}"
        candidate = (base[: max_base - len(tag)] + tag + suffix).strip()
        if candidate not in used:
            used.add(candidate)
            return candidate

    raise ValueError(f"Could not generate unique sheet name for column: {col_name}")


def count_variances(df: pd.DataFrame) -> dict[str, int]:
    counts = {}
    test_cols = [c for c in df.columns if c.endswith("_test")]
    for tc in test_cols:
        base = tc[: -len("_test")]
        pc = f"{base}_prod"
        if pc not in df.columns:
            continue
        differs = ~(
            df[tc].eq(df[pc]) |
            (df[tc].isna() & df[pc].isna())
        )
        count = int(differs.sum())
        if count > 0:
            counts[base] = count
    return counts


def extract_col_variance_rows(
    df: pd.DataFrame, col_name: str, fname: str, stay_id: str
) -> list[dict]:
    """
    Return one dict per row in the variance sheet where col_name differed.
    """
    tc = f"{col_name}_test"
    pc = f"{col_name}_prod"
    if tc not in df.columns or pc not in df.columns:
        return []

    differs = ~(
        df[tc].eq(df[pc]) |
        (df[tc].isna() & df[pc].isna())
    )
    result = []
    for _, row in df[differs].iterrows():
        result.append({
            "File":    fname,
            "Stay ID": stay_id,
            tc:        row[tc],
            pc:        row[pc],
        })
    return result


def process_directory(directory: str) -> tuple[list[dict], dict[str, list[dict]]]:
    """
    Returns:
        summary_rows  : list of dicts for the Variance Summary sheet
        detail_data   : {col_name: [row_dicts, ...]} for each column detail sheet
    """
    pattern = os.path.join(directory, FILE_PATTERN)
    files   = sorted(glob.glob(pattern))

    if not files:
        print(f"No files found matching: {pattern}")
        return [], {}

    summary_rows: list[dict]            = []
    detail_data:  dict[str, list[dict]] = {}

    for fpath in files:
        fname   = os.path.basename(fpath)
        stay_id = extract_stay_id(fname)
        print(f"  Processing: {fname}")

        try:
            xl = pd.read_excel(fpath, sheet_name=None, dtype=str)
        except Exception as e:
            print(f"    ERROR reading file: {e}")
            summary_rows.append({"File": fname, "Stay ID": stay_id, "_error": str(e)})
            continue

        if VARIANCE_SHEET not in xl:
            print(f"    WARNING: no '{VARIANCE_SHEET}' sheet found — skipping")
            summary_rows.append({
                "File": fname, "Stay ID": stay_id,
                "_error": f"No '{VARIANCE_SHEET}' sheet"
            })
            continue

        variance_df = xl[VARIANCE_SHEET]
        counts      = count_variances(variance_df)

        row = {"File": fname, "Stay ID": stay_id}
        row.update(counts)
        summary_rows.append(row)

        for col_name in counts:
            rows = extract_col_variance_rows(variance_df, col_name, fname, stay_id)
            detail_data.setdefault(col_name, []).extend(rows)

    return summary_rows, detail_data


# ── shared style helpers ──────────────────────────────────────────────────────
def _styles() -> dict:
    thin_side = Side(style="thin", color="BFBFBF")
    return {
        "HEADER_FILL":   PatternFill("solid", start_color="1F4E79", end_color="1F4E79"),
        "ALT_FILL":      PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0"),
        "VARIANCE_FILL": PatternFill("solid", start_color="FDEBD0", end_color="FDEBD0"),
        "NO_FILL":       PatternFill(fill_type=None),
        "HEADER_FONT":   Font(name="Arial", bold=True, color="FFFFFF", size=11),
        "BODY_FONT":     Font(name="Arial", size=10),
        "BOLD_FONT":     Font(name="Arial", size=10, bold=True),
        "CENTER":        Alignment(horizontal="center", vertical="center", wrap_text=True),
        "LEFT":          Alignment(horizontal="left",   vertical="center", wrap_text=True),
        "BORDER":        Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side),
    }


def _autofit(ws) -> None:
    for col_cells in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max(max_len + 4, 12), 60
        )


def _style_header_row(ws, s: dict) -> list:
    header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
    for cell in header_row:
        cell.font = s["HEADER_FONT"]
        cell.fill = s["HEADER_FILL"]
        cell.alignment = s["CENTER"]
        cell.border = s["BORDER"]
    return [cell.value for cell in header_row]


def _format_summary_sheet(ws, variance_cols: list[str], s: dict) -> None:
    col_names = _style_header_row(ws, s)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        alt = row_idx % 2 == 0
        for cell in row:
            col_name    = col_names[cell.column - 1]
            is_variance = col_name in variance_cols
            if is_variance and isinstance(cell.value, (int, float)) and cell.value > 0:
                cell.fill = s["VARIANCE_FILL"]
            elif alt:
                cell.fill = s["ALT_FILL"]
            else:
                cell.fill = s["NO_FILL"]
            cell.font      = s["BODY_FONT"]
            cell.alignment = s["CENTER"] if is_variance else s["LEFT"]
            cell.border    = s["BORDER"]

    # Total Variances column
    var_col_indices = [i + 1 for i, name in enumerate(col_names) if name in variance_cols]
    total_col_idx   = ws.max_column + 1

    hdr            = ws.cell(1, total_col_idx, "Total Variances")
    hdr.font       = s["HEADER_FONT"]
    hdr.fill       = s["HEADER_FILL"]
    hdr.alignment  = s["CENTER"]
    hdr.border     = s["BORDER"]

    for row_idx in range(2, ws.max_row + 1):
        val = (
            f"={'+'.join(get_column_letter(c) + str(row_idx) for c in var_col_indices)}"
            if var_col_indices else 0
        )
        cell           = ws.cell(row_idx, total_col_idx, val)
        cell.font      = s["BOLD_FONT"]
        cell.alignment = s["CENTER"]
        cell.border    = s["BORDER"]

    ws.freeze_panes = "A2"
    _autofit(ws)


def _format_detail_sheet(ws, value_cols: set[str], s: dict) -> None:
    """
    value_cols: set of the two paired column names (e.g. col_test, col_prod)
    that contain the actual differing values — styled with VARIANCE_FILL.
    """
    col_names = _style_header_row(ws, s)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        alt = row_idx % 2 == 0
        for cell in row:
            col_name = col_names[cell.column - 1]
            is_value = col_name in value_cols
            if is_value:
                cell.fill = s["VARIANCE_FILL"]
            elif alt:
                cell.fill = s["ALT_FILL"]
            else:
                cell.fill = s["NO_FILL"]
            cell.font      = s["BODY_FONT"]
            cell.alignment = s["CENTER"] if is_value else s["LEFT"]
            cell.border    = s["BORDER"]

    ws.freeze_panes = "A2"
    _autofit(ws)


def build_summary(
    summary_rows: list[dict],
    detail_data:  dict[str, list[dict]],
    output_path:  str,
) -> None:
    summary_df = pd.DataFrame(summary_rows)

    meta_cols     = ["File", "Stay ID", "_error"]
    variance_cols = sorted([c for c in summary_df.columns if c not in meta_cols])

    for col in variance_cols:
        summary_df[col] = pd.to_numeric(summary_df[col], errors="coerce").fillna(0).astype(int)

    final_cols = ["File", "Stay ID"] + variance_cols
    if "_error" in summary_df.columns:
        final_cols.append("_error")
    summary_df = summary_df[final_cols]
    summary_df.rename(columns={"_error": "Error / Notes"}, inplace=True)

    if os.path.exists(output_path):
        os.remove(output_path)

    # Build sheet-name map: col_name -> unique Excel sheet name
    used_sheet_names: set[str] = {"Variance Summary"}
    sheet_name_map: dict[str, str] = {}
    for col in variance_cols:
        sheet_name_map[col] = make_sheet_name(col, used_sheet_names)

    # ── write all sheets via ExcelWriter ─────────────────────────────────────
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Variance Summary")

        for col in variance_cols:
            rows = detail_data.get(col, [])
            df   = pd.DataFrame(rows) if rows else pd.DataFrame(
                columns=["File", "Stay ID", f"{col}_test", f"{col}_prod"]
            )
            df.to_excel(writer, index=False, sheet_name=sheet_name_map[col])

    # ── format all sheets ─────────────────────────────────────────────────────
    wb = load_workbook(output_path)
    s  = _styles()

    _format_summary_sheet(wb["Variance Summary"], variance_cols, s)

    for col in variance_cols:
        ws         = wb[sheet_name_map[col]]
        value_cols = {f"{col}_test", f"{col}_prod"}
        _format_detail_sheet(ws, value_cols, s)

    wb.save(output_path)

    print(f"\nSummary saved to: {output_path}")
    print(f"  {len(summary_rows)} file(s) processed")
    print(f"  {len(variance_cols)} column(s) with variances — detail sheet created for each:")
    for col in variance_cols:
        row_count = len(detail_data.get(col, []))
        print(f"    '{sheet_name_map[col]}'  ({row_count} row(s))")


def main():
    file_dir    = r'C:\Users\jwalker221\OneDrive - Hilton\Jira\Stay Derived\orig_stay_booking_by_stay_date'
    output_path = r'C:\Users\jwalker221\OneDrive - Hilton\Jira\Stay Derived\orig_stay_booking_by_stay_date\orig_stay_booking_by_stay_date test environment vs PROD existing (original specific) summary results.xlsx'
    directory   = sys.argv[1] if len(sys.argv) > 1 else file_dir
    output_path = sys.argv[2] if len(sys.argv) > 2 else output_path

    print(f"Scanning: {os.path.abspath(directory)}")
    summary_rows, detail_data = process_directory(directory)

    if not summary_rows:
        print("Nothing to summarize.")
        return

    build_summary(summary_rows, detail_data, output_path)


if __name__ == "__main__":
    main()

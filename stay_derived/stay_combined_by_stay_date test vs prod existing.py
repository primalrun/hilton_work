"""
compare_test_vs_prod.py

Compares TEST (dw_preprod) vs PROD (dw_prod) Redshift environments for a random
sample of stays, using the SOURCE/TARGET/VARIANCE pattern.

For each sampled (stay_id, prop_cd):
    1. Runs the TEST dynamic SELECT against dw_preprod
    2. Runs the PROD dynamic SELECT against dw_prod
    3. Compares column-by-column
    4. Writes a formatted, per-stay .xlsx workbook to PROJECT_DIR

Requirements:
    pip install psycopg2-binary pandas openpyxl

Author: generated for Jason Walker
"""

import configparser
import logging
import os
import sys
import random
import warnings
from datetime import datetime

import pandas as pd
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings(
    "ignore",
    message="pandas only supports SQLAlchemy connectable.*",
    category=UserWarning,
)

# --------------------------------------------------------------------------
# CONFIG -- edit these for your environment
# --------------------------------------------------------------------------

PROJECT_DIR = r"C:\Users\jwalker221\OneDrive - Hilton\Jira\Stay Derived\stay_combined_by_stay_date"
INI_PATH = r"C:\Users\jwalker221\OneDrive - Hilton\Documents\cred.ini"

PROD_SECTION = "dw_prod"
TEST_SECTION = "dw_preprod"

TEST_TABLE = "stay_combined_by_stay_date"

# SQL file names -- must live in PROJECT_DIR
POPULATION_SQL_FILENAME = "random stay_id TEST vs. PROD EXISTING that exists in TEST.sql"
TEST_SELECT_SQL_FILENAME = "TEST vs. PROD EXISTING (TEST SELECT dynamic).sql"
PROD_SELECT_SQL_FILENAME = "TEST vs. PROD EXISTING (PROD EXISTING SELECT dynamic).sql"

# Token in the dynamic SQL files to be replaced with the actual stay_id per loop iteration
STAY_ID_PLACEHOLDER = "stay_id_variable"
# Optional -- replaced only if present in the SQL file text
PROP_CD_PLACEHOLDER = "prop_cd_variable"

# How many stays to pull from the population query, and whether to subsample further in Python
SAMPLE_SIZE = None  # set an int to randomly subsample the population query results in Python; None = use all rows returned

# Decimal precision for numeric comparisons (mirrors the round(10) fix used on
# curr_stay_booking_by_stay_date to avoid float noise flagging false variances)
ROUND_DECIMALS = 10

LOG_LEVEL = logging.INFO

# --------------------------------------------------------------------------

logging.basicConfig(
    level=LOG_LEVEL,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


def load_env_config(ini_path: str, section: str) -> dict:
    """Read one [section] from the credentials .ini file."""
    parser = configparser.ConfigParser()
    read_files = parser.read(ini_path)
    if not read_files:
        raise FileNotFoundError(f"Could not read ini file at: {ini_path}")
    if section not in parser:
        raise KeyError(
            f"Section [{section}] not found in {ini_path}. "
            f"Available sections: {parser.sections()}"
        )
    sect = parser[section]
    return {
        "host": sect.get("host"),
        "port": sect.getint("port", fallback=5439),
        "dbname": sect.get("dbname"),
        "user": sect.get("user"),
        "password": sect.get("password"),
    }


def get_connection(env_cfg: dict):
    """
    Open a raw psycopg2 connection to Redshift. Note: we deliberately do NOT
    use a SQLAlchemy engine here. SQLAlchemy's postgres dialect tries to SET
    standard_conforming_strings during connection init, which Redshift's
    Postgres fork rejects with 'unrecognized configuration parameter'. Plain
    psycopg2 avoids that entirely -- pandas' read_sql works fine with it, it
    just emits a cosmetic UserWarning, which we suppress below.
    """
    return psycopg2.connect(
        host=env_cfg["host"],
        port=env_cfg["port"],
        dbname=env_cfg["dbname"],
        user=env_cfg["user"],
        password=env_cfg["password"],
        connect_timeout=30,
    )


def read_sql_file(project_dir: str, filename: str) -> str:
    path = os.path.join(project_dir, filename)
    if not os.path.exists(path):
        raise FileNotFoundError(f"SQL file not found: {path}")
    with open(path, "r", encoding="utf-8-sig") as f:
        return f.read()


def get_population(conn, sql_text: str, sample_size=None) -> list:
    """
    Runs the population SQL against the TEST connection and returns a list of
    dicts: [{"stay_id": ..., "prop_cd": ...}, ...]
    """
    df = pd.read_sql(sql_text, conn)
    df.columns = [c.lower() for c in df.columns]

    if "stay_id" not in df.columns or "prop_cd" not in df.columns:
        raise ValueError(
            f"Population query must return stay_id and prop_cd columns. "
            f"Got columns: {list(df.columns)}"
        )

    records = df[["stay_id", "prop_cd"]].drop_duplicates().to_dict("records")

    if sample_size and sample_size < len(records):
        records = random.sample(records, sample_size)

    log.info("Population query returned %d unique (stay_id, prop_cd) pairs", len(records))
    return records


def build_dynamic_sql(sql_template: str, stay_id, prop_cd) -> str:
    sql = sql_template.replace(STAY_ID_PLACEHOLDER, str(stay_id))
    if PROP_CD_PLACEHOLDER in sql:
        # wrap in quotes in case prop_cd is a string code like 'ATL01'
        sql = sql.replace(PROP_CD_PLACEHOLDER, str(prop_cd))
    return sql


def run_query(conn, sql_text: str) -> pd.DataFrame:
    df = pd.read_sql(sql_text, conn)
    df.columns = [c.lower() for c in df.columns]
    return df


def _round_if_numeric(val):
    try:
        return round(float(val), ROUND_DECIMALS)
    except (TypeError, ValueError):
        return val


def compare_dataframes(df_test: pd.DataFrame, df_prod: pd.DataFrame) -> pd.DataFrame:
    """
    Row-by-row AND column-by-column comparison is purely positional: TEST and
    PROD SQL are written to return columns in the same order, so column index
    is used to pair them up rather than matching on column name. This still
    works if a column happens to be labeled slightly differently between the
    two queries, as long as the underlying order lines up.

    Returns a tidy comparison DataFrame:
    row_num | column_position | column_name | test_value | prod_value | variance | match
    """
    rows = []

    test_cols = list(df_test.columns)
    prod_cols = list(df_prod.columns)
    max_cols = max(len(test_cols), len(prod_cols))
    max_rows = max(len(df_test), len(df_prod))

    if len(df_test) != len(df_prod):
        rows.append(
            {
                "row_num": "ALL",
                "column_position": "",
                "column_name": "__ROW_COUNT__",
                "test_value": len(df_test),
                "prod_value": len(df_prod),
                "variance": len(df_test) - len(df_prod),
                "match": "N",
            }
        )

    if len(test_cols) != len(prod_cols):
        rows.append(
            {
                "row_num": "ALL",
                "column_position": "",
                "column_name": "__COLUMN_COUNT__",
                "test_value": len(test_cols),
                "prod_value": len(prod_cols),
                "variance": len(test_cols) - len(prod_cols),
                "match": "N",
            }
        )

    for row_idx in range(max_rows):
        for col_idx in range(max_cols):
            test_col_name = test_cols[col_idx] if col_idx < len(test_cols) else None
            prod_col_name = prod_cols[col_idx] if col_idx < len(prod_cols) else None

            # label the column using whichever name is available; note both if they differ
            if test_col_name and prod_col_name and test_col_name != prod_col_name:
                col_label = f"{test_col_name} / {prod_col_name}"
            else:
                col_label = test_col_name or prod_col_name

            if test_col_name is None:
                test_val = "COLUMN MISSING IN TEST"
            elif row_idx < len(df_test):
                test_val = df_test.iloc[row_idx, col_idx]
            else:
                test_val = "ROW MISSING IN TEST"

            if prod_col_name is None:
                prod_val = "COLUMN MISSING IN PROD"
            elif row_idx < len(df_prod):
                prod_val = df_prod.iloc[row_idx, col_idx]
            else:
                prod_val = "ROW MISSING IN PROD"

            test_rounded = _round_if_numeric(test_val)
            prod_rounded = _round_if_numeric(prod_val)

            is_numeric = isinstance(test_rounded, float) and isinstance(prod_rounded, float)

            if is_numeric:
                variance = round(test_rounded - prod_rounded, ROUND_DECIMALS)
                match = "Y" if variance == 0 else "N"
            else:
                variance = ""
                match = "Y" if str(test_val) == str(prod_val) else "N"

            rows.append(
                {
                    "row_num": row_idx + 1,
                    "column_position": col_idx + 1,
                    "column_name": col_label,
                    "test_value": test_val,
                    "prod_value": prod_val,
                    "variance": variance,
                    "match": match,
                }
            )

    return pd.DataFrame(rows)


def write_excel(comparison_df: pd.DataFrame, out_path: str, stay_id, prop_cd, test_table: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    body_font = Font(name="Calibri")
    mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin = Side(border_style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- summary block ---
    ws["A1"] = "Table"
    ws["B1"] = test_table
    ws["A2"] = "stay_id"
    ws["B2"] = stay_id
    ws["A3"] = "prop_cd"
    ws["B3"] = prop_cd
    ws["A4"] = "Run timestamp"
    ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    total_checks = len(comparison_df)
    mismatches = len(comparison_df[comparison_df["match"] == "N"])
    ws["A5"] = "Total checks"
    ws["B5"] = total_checks
    ws["A6"] = "Mismatches"
    ws["B6"] = mismatches

    for r in range(1, 7):
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"].font = body_font

    header_row = 8
    headers = ["row_num", "column_position", "column_name", "test_value", "prod_value", "variance", "match"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header.upper())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for i, record in enumerate(comparison_df.to_dict("records"), start=header_row + 1):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=i, column=col_idx, value=record[header])
            cell.font = body_font
            cell.border = border
            if header == "match":
                cell.alignment = Alignment(horizontal="center")
            if record["match"] == "N":
                cell.fill = mismatch_fill

    ws.freeze_panes = f"A{header_row + 1}"

    # approximate autofit
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for record in comparison_df.to_dict("records"):
            val_len = len(str(record[header]))
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    wb.save(out_path)


def main():
    log.info("Loading Redshift connection config from %s", INI_PATH)
    test_cfg = load_env_config(INI_PATH, TEST_SECTION)
    prod_cfg = load_env_config(INI_PATH, PROD_SECTION)

    log.info("Connecting to TEST (%s) and PROD (%s)", TEST_SECTION, PROD_SECTION)
    test_conn = get_connection(test_cfg)
    prod_conn = get_connection(prod_cfg)

    summary_log = []

    try:
        # --- Step 1: population sample from TEST ---
        population_sql = read_sql_file(PROJECT_DIR, POPULATION_SQL_FILENAME)
        population = get_population(test_conn, population_sql, SAMPLE_SIZE)

        test_select_template = read_sql_file(PROJECT_DIR, TEST_SELECT_SQL_FILENAME)
        prod_select_template = read_sql_file(PROJECT_DIR, PROD_SELECT_SQL_FILENAME)

        # --- Step 2: loop over population ---
        for i, record in enumerate(population, start=1):
            stay_id = record["stay_id"]
            prop_cd = record["prop_cd"]
            log.info("[%d/%d] Comparing stay_id=%s prop_cd=%s", i, len(population), stay_id, prop_cd)

            try:
                test_sql = build_dynamic_sql(test_select_template, stay_id, prop_cd)
                prod_sql = build_dynamic_sql(prod_select_template, stay_id, prop_cd)

                df_test = run_query(test_conn, test_sql)
                df_prod = run_query(prod_conn, prod_sql)

                comparison_df = compare_dataframes(df_test, df_prod)

                out_path = os.path.join(
                    PROJECT_DIR,
                    f"{TEST_TABLE} test environment vs PROD existing ({prop_cd} stay_id {stay_id}).xlsx",
                )
                write_excel(comparison_df, out_path, stay_id, prop_cd, TEST_TABLE)

                mismatches = len(comparison_df[comparison_df["match"] == "N"])
                summary_log.append(
                    {"stay_id": stay_id, "prop_cd": prop_cd, "status": "OK", "mismatches": mismatches}
                )
                log.info("  -> wrote %s (%d mismatches)", os.path.basename(out_path), mismatches)

            except Exception as exc:
                log.error("  -> FAILED for stay_id=%s prop_cd=%s: %s", stay_id, prop_cd, exc)
                summary_log.append(
                    {"stay_id": stay_id, "prop_cd": prop_cd, "status": f"ERROR: {exc}", "mismatches": ""}
                )
                continue

    finally:
        test_conn.close()
        prod_conn.close()
        log.info("Closed TEST and PROD connections")

    # write a run summary CSV alongside the per-stay workbooks
    summary_path = os.path.join(
        PROJECT_DIR, f"{TEST_TABLE} run summary {datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    )
    pd.DataFrame(summary_log).to_csv(summary_path, index=False)
    log.info("Run summary written to %s", summary_path)


if __name__ == "__main__":
    main()

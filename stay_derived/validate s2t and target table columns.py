import pandas as pd
import os
import openpyxl
from openpyxl.utils import get_column_letter
import configparser
import redshift_connector as rc
import sys
import numpy as np

project_dir = r'C:\Users\jwalker221\OneDrive - Hilton\Jira\Stay Derived\stay_combined_by_stay_date'
file_config = r'C:\Users\jwalker221\OneDrive - Hilton\Documents\cred.ini'
file_s2t_name = 'S2T_StayDerived_CombinedStayDate.xlsx'
st2_sheet_name = 'S2T'
target_schema = 'bdp_stay'
target_table = 'stay_combined_by_stay_date'
file_temp_name = 'temp.xlsx'

file_s2t = os.path.join(project_dir, file_s2t_name)
file_result = os.path.join(project_dir, f'{target_table} column and data type validation.xlsx')
file_temp = os.path.join(project_dir, file_temp_name)


# read s2t file, get columns and data type
wb = openpyxl.load_workbook(file_s2t)

try:
    ws = wb[st2_sheet_name]
except KeyError:
    print(f'sheet {st2_sheet_name} not found in {file_s2t}')
    print('process cancelled')
    sys.exit()

search_range = ws['B1:B50']
found_cell = None

for row in search_range:
    for cell in row:
        if cell.value == 'SOURCE':
            found_cell = cell
            break
        if found_cell:
            break

if found_cell:
    header_row = found_cell.row
else:
    print(f'did not find SOURCE in column B')
    print('process cancelled')
    sys.exit()

search_range_str = f'B{header_row}:AZ{header_row}'
search_range = ws[search_range_str]
found_cell = None

for row in search_range:
    for cell in row:
        if cell.value == 'TARGET':
            found_cell = cell
            break
        if found_cell:
            break

header_row = header_row + 1
header_col = found_cell.column

start_col = header_col
start_row = header_row
end_col = start_col

while ws.cell(row=start_row,column=end_col).value is not None:
    end_col += 1

# back up one column to get last data column
end_col = end_col - 1

end_row = start_row

while ws.cell(row=end_row,column=start_col + 1).value is not None:
    end_row += 1

# back up one row to get last data row
end_row = end_row - 1

col_range = f'{get_column_letter(start_col)}:{get_column_letter(end_col)}'


df_s2t = pd.read_excel(
    file_s2t
    , sheet_name=st2_sheet_name
    ,skiprows=start_row - 1
    ,nrows=end_row - start_row + 1
    ,usecols=col_range
)

df_s2t.columns = [c.split('.')[0] for c in df_s2t.columns]

column_raname_dict = {
    'Net New': 'net_new'
    ,'Schema': 'schema'
    ,'Table/View': 'object'
    ,'Column': 'column_name'
    ,'Data Type': 'data_type'
    ,'Length': 'length'
}

# remove Net New if not exist in S2T
if 'Net New' not in df_s2t.columns:
    del column_raname_dict['Net New']

df_s2t = df_s2t.rename(columns = column_raname_dict)
df_s2t = df_s2t[list(column_raname_dict.values())]

config = configparser.ConfigParser()
config.read(file_config)
dw_config = config['dw_preprod']
host = dw_config['host']
dbname = dw_config['dbname']
user = dw_config['user']
password = dw_config['password']

rs_cred = [
    host
    ,dbname
    ,user
    ,password
]


def connect_rs(host_p, database_p, user_p, password_p):
    conn = rc.connect(
        host=host_p
        , database=database_p
        , user=user_p
        , password=password_p
    )
    conn.autocommit = True
    return conn

def data_type_adjusted(data_type):
    if data_type == 'character varying':
        return 'varchar'
    elif data_type == 'timestamp without time zone':
        return 'timestamp'
    else:
        return data_type


def validation_column(col_s2t, col_target):
    if col_s2t is None and col_target is None:
        return 'good'
    elif col_s2t is not None and col_target is None:
        return 'bad'
    elif col_s2t is None and col_target is not None:
        return 'bad'
    elif col_s2t == col_target:
        return 'good'
    else:
        return 'bad'


def clean_excel_file(wb_in, file_excel_p):
    for sheet in wb_in.sheetnames:
        if str(sheet)[0:5] == 'Sheet':
            wb_in.remove(wb_in[sheet])

    for sheet in wb_in.sheetnames:
        ws = wb_in[sheet]
        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max([len(str(cell.value) or "") for cell in column_cells])
            max_length = max_length + 3
            ws.column_dimensions[column_letter].width = max_length
            ws.freeze_panes = 'A2'
    wb_in.save(file_excel_p)


def delete_file_if_exists(file_path_p):
    if os.path.exists(file_path_p):
        os.remove(file_path_p)


def total_validation(object_validation, length_validation, data_type_validation):
    all_good = all(x == 'good' for x in [object_validation, length_validation, data_type_validation])
    if all_good:
        return 'good'
    else:
        return 'bad'




rs_host, rs_database, rs_user, rs_pword = rs_cred
rs_conn = connect_rs(rs_host, rs_database, rs_user, rs_pword)
cursor = rs_conn.cursor()

sql = f"""
select table_schema as schema,
    table_name as object, 
    column_name, 
    data_type, 
    character_maximum_length as length   
FROM information_schema.columns 
WHERE 1 = 1
	and table_schema = '{target_schema}'
	and table_name = '{target_table}'
ORDER BY ordinal_position;
"""

cursor.execute(sql)
df_target = cursor.fetch_dataframe()
df_target['length'] = df_target['length'].astype('Int32')





df_compare = df_s2t.merge(df_target, how='outer', on='column_name', suffixes=('_S2T', '_target'))
df_compare['data_type_target_adjusted'] = df_compare.apply(lambda x: data_type_adjusted(x['data_type_target']), axis=1)

missing_col = ['object_S2T', 'object_target', 'length_S2T', 'length_target', 'data_type_S2T', 'data_type_target_adjusted']
df_compare['object_S2T'] = df_compare['object_S2T'].fillna('')
df_compare['object_target'] = df_compare['object_target'].fillna('')
df_compare['length_S2T'] = df_compare['length_S2T'].fillna(999)
df_compare['length_target'] = df_compare['length_target'].fillna(999)
df_compare['data_type_S2T'] = df_compare['data_type_S2T'].fillna('')
df_compare['data_type_target'] = df_compare['data_type_target'].fillna('')

df_compare['object_validation'] = df_compare.apply(lambda x: validation_column(x['object_S2T'], x['object_target']), axis=1)
df_compare['length_validation'] = df_compare.apply(lambda x: validation_column(x['length_S2T'], x['length_target']), axis=1)
df_compare['data_type_validation'] = df_compare.apply(lambda x: validation_column(x['data_type_S2T'], x['data_type_target_adjusted']), axis=1)

df_compare['length_S2T'] = df_compare['length_S2T'].replace(999, np.nan)
df_compare['length_target'] = df_compare['length_target'].replace(999, np.nan)


df_compare['total_validation'] = df_compare.apply(
    lambda x: total_validation(x['object_validation'], x['length_validation'], x['data_type_validation']), axis=1)

column_order = [
    'total_validation'
    ,'net_new'
    ,'schema_S2T'
    ,'schema_target'
    ,'object_S2T'
    ,'object_target'
    ,'object_validation'
    ,'column_name'
    ,'length_S2T'
    ,'length_target'
    ,'length_validation'
    ,'data_type_S2T'
    ,'data_type_target'
    ,'data_type_target_adjusted'
    ,'data_type_validation'
]

# remove Net New if not exist in S2T
if 'Net New' not in df_s2t.columns:
    column_order = [col for col in column_order if col != 'net_new']

df_compare = df_compare[column_order]

delete_file_if_exists(file_result)
df_compare.to_excel(file_result, index=False, sheet_name='validation')

#clean excel file
wb = openpyxl.load_workbook(filename=file_result)
clean_excel_file(wb, file_result)
wb.close()

os.startfile(file_result)
print('success')


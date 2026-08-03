import pandas as pd
import os
import openpyxl
from openpyxl.utils import get_column_letter
import configparser
import redshift_connector as rc
import sys
import numpy as np


project_dir = r'C:\Users\jwalker221\OneDrive - Hilton\Jira\Stay Derived\curr_stay_booking_by_stay_date'
file_config = r'C:\Users\jwalker221\OneDrive - Hilton\Documents\cred.ini'
target_schema = 'bdp_stay'
test_table = 'curr_stay_booking_by_stay_date'
prod_table = 'f_stay_booking_summary'
file_temp_name = 'temp.xlsx'
col_map_old_to_new_name = 'column mapping existing to new'
file_sql_random_stay_name = 'random property and stay population.sql'
file_s2t_name = 'curr_stay_booking_by_stay_date column and data type validation 20260526 1507.xlsx'

file_result = os.path.join(project_dir, f'{test_table} test environment vs PROD existing.xlsx')
file_temp = os.path.join(project_dir, file_temp_name)
file_col_map_old_to_new = os.path.join(project_dir, f'{col_map_old_to_new_name}.xlsx')
file_sql_random_stay = os.path.join(project_dir, file_sql_random_stay_name)
file_s2t = os.path.join(project_dir, file_s2t_name)


def connect_rs(host_p, database_p, user_p, password_p):
    conn = rc.connect(
        host=host_p
        , database=database_p
        , user=user_p
        , password=password_p
    )
    conn.autocommit = True
    return conn


def write_df_to_excel_temp_file(df, file_name, sheet_name='temp'):
    if os.path.exists(file_name):
        os.remove(file_name)
    df.to_excel(file_name, index=False, sheet_name=sheet_name)


def create_excel_workbook(file_name_in):
    wb = openpyxl.Workbook()
    wb.save(file_name_in)
    wb.close()


def delete_file_if_exists(file_path_p):
    if os.path.exists(file_path_p):
        os.remove(file_path_p)


def append_df_to_existing_excel_workbook(df_p, workbook_path_p, sheet_name_p):
    with pd.ExcelWriter(workbook_path_p, mode='a', engine='openpyxl') as writer:
        df_p.to_excel(writer, sheet_name=sheet_name_p, index=False)


def clean_excel_file(wb_in, file_excel_p):
    for sheet in wb_in.sheetnames:
        if str(sheet)[0:5] == 'Sheet':
            wb_in.remove(wb_in[sheet])

    for sheet in wb_in.sheetnames:
        ws = wb_in[sheet]
        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max([len(str(cell.value) or "") for cell in column_cells])
            max_length = max_length + 3
            ws.column_dimensions[column_letter].width = max_length
            ws.freeze_panes = 'A2'
    wb_in.save(file_excel_p)


def col_variance_reason(col_name, exists_test, exists_prod, variance):
    if variance == 'no':
        return None
    if (exists_prod == 'yes' and exists_test != 'yes'):
        return 'in prod, not in test'
    if (exists_test == 'yes' and exists_prod != 'yes'):
        return 'in test, not in prod'


def column_name_prod(column_name, existing_column_name_prod, variance, exists_prod):
    if exists_prod == 'no':
        return None
    elif (exists_prod == 'yes') and (variance == 'no'):
        if existing_column_name_prod:
            return existing_column_name_prod
        else:
            return column_name


def convert_data_type(data_type):
    if data_type in ['integer', 'bigint']:
        return 'Int64'
    elif data_type in ['character varying', 'date', 'timestamp without time zone']:
        return 'string'
    elif data_type == 'boolean':
        return 'boolean'
    elif data_type == 'double precision':
        return 'float'
    else:
        return 'string'



config = configparser.ConfigParser()
config.read(file_config)
dw_config_test = config['dw_preprod']
host = dw_config_test['host']
dbname = dw_config_test['dbname']
user = dw_config_test['user']
password = dw_config_test['password']

rs_cred = [
    host
    ,dbname
    ,user
    ,password
]

rs_host, rs_database, rs_user, rs_pword = rs_cred
rs_conn_test = connect_rs(rs_host, rs_database, rs_user, rs_pword)
cursor_test = rs_conn_test.cursor()

sql = f"""
select *
from {target_schema}.{test_table}
limit 1
"""
cursor_test.execute(sql)
df_test = cursor_test.fetch_dataframe()
columns_test = df_test.columns.values.tolist()


dw_config_prod = config['dw_prod']
host = dw_config_prod['host']
dbname = dw_config_prod['dbname']
user = dw_config_prod['user']
password = dw_config_prod['password']

rs_cred = [
    host
    ,dbname
    ,user
    ,password
]

rs_host, rs_database, rs_user, rs_pword = rs_cred
rs_conn_prod = connect_rs(rs_host, rs_database, rs_user, rs_pword)
cursor_prod = rs_conn_prod.cursor()

sql = f"""
select *
from {target_schema}.{prod_table}
limit 1
"""

cursor_prod.execute(sql)
df_prod = cursor_prod.fetch_dataframe()
columns_prod = df_prod.columns.values.tolist()

delete_file_if_exists(file_result)
create_excel_workbook(file_result)

columns_test_set = set(columns_test)

df_col_map_old_to_new = pd.read_excel(file_col_map_old_to_new)
df_col_map_old_to_new = df_col_map_old_to_new[df_col_map_old_to_new['new'] != '*no mapping*']
df_col_map_old_to_new_no_mapping = df_col_map_old_to_new[df_col_map_old_to_new['new'] == '*no mapping*']


col_map_old_to_new_list = df_col_map_old_to_new.to_numpy().tolist()
dict_old_to_new = {k: v for k, v in col_map_old_to_new_list}
dict_new_to_old = {v: k for k, v in dict_old_to_new.items()}

col_map_old_to_new_no_mapping_list = df_col_map_old_to_new_no_mapping.to_numpy().tolist()
dict_old_to_new_no_mapping = {k: v for k, v in col_map_old_to_new_no_mapping_list}

columns_prod = [dict_old_to_new[col] if col in dict_old_to_new else col for col in columns_prod]
columns_prod_set = set(columns_prod)

columns_match = list(columns_test_set.intersection(columns_prod_set))
columns_test_only = list(columns_test_set.difference(columns_prod_set))
columns_prod_only = list(columns_prod_set.difference(columns_test_set))

df_columns_match = pd.DataFrame(data=columns_match, columns=['column_name'])
append_df_to_existing_excel_workbook(df_columns_match, file_result, 'column match')

df_columns_test_only = pd.DataFrame(data=columns_test_only, columns=['column_name'])
append_df_to_existing_excel_workbook(df_columns_test_only, file_result, 'column test only')

df_columns_prod_only = pd.DataFrame(data=columns_prod_only, columns=['column_name'])
append_df_to_existing_excel_workbook(df_columns_prod_only, file_result, 'column prod only')

df_columns_test = pd.DataFrame(data=columns_test, columns=['column_name'])
df_columns_test['exists'] = 'yes'
df_columns_prod = pd.DataFrame(data=columns_prod, columns=['column_name'])
df_columns_prod['exists'] = 'yes'
df_columns_all = df_columns_test.merge(df_columns_prod, on='column_name', how='outer', suffixes=('_test', '_prod'))
df_columns_all['variance'] = df_columns_all.apply(
    lambda x: 'no' if x['exists_test'] == x['exists_prod'] else 'yes', axis=1)
df_columns_all['existing column_name prod'] = df_columns_all.apply(
    lambda x: dict_new_to_old[x['column_name']]
    if x['column_name'] in dict_new_to_old
    else None, axis=1)
df_columns_all['variance reason'] = df_columns_all.apply(
    lambda x: col_variance_reason(x['column_name'], x['exists_test'], x['exists_prod'], x['variance']), axis=1)
df_columns_all['column name prod'] = df_columns_all.apply(
    lambda x: column_name_prod(
        x['column_name'], x['existing column_name prod'], x['variance'], x['exists_prod']), axis=1
)

append_df_to_existing_excel_workbook(df_columns_all, file_result, 'column comparison')

df_column_compare = df_columns_all[
    (df_columns_all['column name prod'].notnull()) &
    (~df_columns_all['column_name'].str.startswith('dw_'))
][['column_name', 'column name prod']]

column_compare_list = df_column_compare.to_numpy().tolist()
test_columns, prod_columns = map(list, zip(*column_compare_list))
test_agg_column_set = {'curr_booking_room_rate_local_amt'}
test_columns_select = [col if col not in test_agg_column_set else f'sum({col}) as {col}' for col in test_columns]
test_columns_group_by = [col for col in test_columns if col not in test_agg_column_set]
test_columns_select_str = ', '.join(test_columns_select)
test_columns_group_by_str = ', '.join(test_columns_group_by)
prod_columns_select_str = ', '.join(prod_columns)

df_s2t = pd.read_excel(file_s2t)
df_s2t = df_s2t[df_s2t['data_type_target'].notnull()][['column_name', 'data_type_target']]
df_s2t_value_list = df_s2t.to_numpy().tolist()
column_data_type_dict = {row[0]: row[1] for row in df_s2t_value_list}
df_test_columns = pd.DataFrame(data=test_columns, columns=['column_name'])

df_test_columns['data_type_redshift'] = df_test_columns.apply(
    lambda x: column_data_type_dict[x['column_name']], axis=1
)

df_test_columns['data_type_new'] = df_test_columns.apply(
    lambda x: convert_data_type(x['data_type_redshift']), axis=1
)

column_data_type_new = df_test_columns['data_type_new'].to_numpy().tolist()

# ['integer', 'character varying', 'date', 'boolean', 'double precision', 'timestamp without time zone', 'bigint']


#clean excel file
wb = openpyxl.load_workbook(filename=file_result)
clean_excel_file(wb, file_result)
wb.close()


# get random property stays
with open(file_sql_random_stay, 'r') as file_r:
    sql_random_stay = file_r.read()

cursor_test.execute(sql_random_stay)
df_stay = cursor_test.fetch_dataframe()
prop_and_stay = df_stay.to_numpy().tolist()


# loop through each random stay
for elem in prop_and_stay:
    prop_cd = elem[0]
    stay_id = elem[1]
    print(f'processing {prop_cd} stay {stay_id}')

    file_result = os.path.join(project_dir, f'{test_table} test environment vs PROD existing ({prop_cd} stay_id {stay_id}).xlsx')

    test_sql = f"""
    select {test_columns_select_str}
    from {target_schema}.{test_table}
    where stay_id = {stay_id}
    group by {test_columns_group_by_str}
    """

    prod_sql = f"""
    select {prod_columns_select_str}
    from {target_schema}.{prod_table}
    where stay_id = {stay_id}
    """

    cursor_test.execute(test_sql)
    df_test = cursor_test.fetch_dataframe()

    cursor_prod.execute(prod_sql)
    df_prod = cursor_prod.fetch_dataframe()
    # rename prod columns to match test columns
    df_prod.columns = df_test.columns

    df_test_col_types = dict(zip(df_test.columns, column_data_type_new))
    df_prod_col_types = dict(zip(df_prod.columns, column_data_type_new))
    df_test = df_test.astype(df_test_col_types)
    df_prod = df_prod.astype(df_prod_col_types)

    # round float columns before comparison
    float_cols = [col for col, dtype in df_test_col_types.items() if dtype == 'float']
    df_test[float_cols] = df_test[float_cols].round(10)
    df_prod[float_cols] = df_prod[float_cols].round(10)

    df_compare = df_test.compare(df_prod, result_names=('test', 'prod'))

    if not df_compare.empty:
        # Join the MultiIndex column names with an underscore
        df_compare.columns = ['_'.join(col).strip() for col in df_compare.columns.values]

    delete_file_if_exists(file_result)
    create_excel_workbook(file_result)

    append_df_to_existing_excel_workbook(df_test, file_result, 'test')
    append_df_to_existing_excel_workbook(df_prod, file_result, 'prod')

    if not df_compare.empty:
        append_df_to_existing_excel_workbook(df_compare, file_result, 'variance')

    # clean excel file
    wb = openpyxl.load_workbook(filename=file_result)
    clean_excel_file(wb, file_result)
    wb.close()



print('success')

import pandas as pd
import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter
import time

file_select_result = r'C:\Users\jwalker221\OneDrive - Hilton\Jira\Stay Derived\Honors Redemption and Rate Plan\preprod_validation\honors_redemption_rate select extract.xlsx'
file_table_result = r'C:\Users\jwalker221\OneDrive - Hilton\Jira\Stay Derived\Honors Redemption and Rate Plan\preprod_validation\honors_redemption_rate table extract.xlsx'
file_result = r'C:\Users\jwalker221\OneDrive - Hilton\Jira\Stay Derived\Honors Redemption and Rate Plan\preprod_validation\honors_redemption_rate comparison result.xlsx'

def create_excel_workbook(file_name_in):
    wb = openpyxl.Workbook()
    wb.save(file_name_in)
    wb.close()


def delete_file_if_exists(file_path_p):
    if os.path.exists(file_path_p):
        os.remove(file_path_p)


def append_df_to_existing_excel_workbook(df_p, workbook_path_p, sheet_name_p):
    with pd.ExcelWriter(workbook_path_p, mode='a', engine='openpyxl') as writer:
        df_p.to_excel(writer, sheet_name=sheet_name_p, index=False)


def format_excel_data(wb_p, sheet_format_dict_p, file_excel_p):
    for sheet in sheet_format_dict_p.keys():
        ws = wb_p[sheet]
        format_dict = sheet_format_dict_p[sheet][0]
        row_start, row_end = sheet_format_dict_p[sheet][1:]
        if row_start is None:
            row_start = 2
        if row_end is None:
            row_end = ws.max_row
        for fmt in format_dict:
            for c in format_dict[fmt]:
                for r in range(row_start, row_end + 1):
                    ws.cell(row=r, column=c).number_format = fmt
    wb_p.save(file_excel_p)


def clean_excel_file(wb_in, file_excel_p):
    for sheet in wb_in.sheetnames:
        if str(sheet)[0:5] == 'Sheet':
            wb_in.remove(wb_in[sheet])

    for sheet in wb_in.sheetnames:
        ws = wb_in[sheet]
        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max([len(str(cell.value) or "") for cell in column_cells])
            max_length = max_length + 3
            ws.column_dimensions[column_letter].width = max_length
            ws.freeze_panes = 'A2'
    wb_in.save(file_excel_p)


def is_same_text(text1, text2):
    if pd.isna(text1) is True and pd.isna(text2) is True:
        return None
    elif pd.isna(text1) is True or pd.isna(text2) is True:
        return 'variance'
    elif str(text1).lower() == str(text2).lower():
        return None
    else:
        return None


def calc_variance_pct(from_amt, to_amt):
    if pd.isna(from_amt) is True and pd.isna(to_amt) is True:
        return None
    elif pd.isna(from_amt) is True and pd.isna(to_amt) is False:
        return to_amt * -1
    elif pd.isna(to_amt) is True and pd.isna(from_amt) is False:
        return -1
    elif from_amt == 0 and to_amt == 0:
        return 0
    elif from_amt == 0 and to_amt != 0:
        return 1
    else:
        return (to_amt - from_amt)/from_amt




delete_file_if_exists(file_result)
create_excel_workbook(file_result)

sheet_format_dict = {}
# {sheet: [{format: [column number]}, row_start, row_end]}

select_df = pd.read_excel(file_select_result)
table_df = pd.read_excel(file_table_result)

grain_columns = [
    'brand_cd'
    ,'brand_nm'
    ,'rate_plan_cd'
    ,'op_area_level1_desc'
    ,'country_desc'
    ]

# check if unique columns exists in both data sets
select_cols = select_df.columns.tolist()
table_cols = table_df.columns.tolist()
all_exist_select = all(item in select_cols for item in grain_columns)
all_exist_table = all(item in table_cols for item in grain_columns)
if all_exist_select and all_exist_table:
    print('All unique columns exist in both sources')
else:
    print('unique columns are missing')
    print('process cancelled')
    sys.exit()

# check if other columns have same name
select_columns_other = list(set(select_cols) - set(grain_columns))
table_columns_other = list(set(table_cols) - set(grain_columns))
all_other_columns_same = all(item in select_columns_other for item in table_columns_other)
if all_other_columns_same:
    print('all other columns are same in both sources')
else:
    print('other columns are missing or named different')
    print('process cancelled')
    sys.exit()

compare_df = select_df.merge(table_df, how='outer', on=grain_columns, suffixes=('_select', '_table'))

non_grain_columns = [col for col in select_columns_other]
non_grain_columns = [f'{col}_select' for col in non_grain_columns] + [f'{col}_table' for col in non_grain_columns]
non_grain_columns = sorted(non_grain_columns)
column_order = grain_columns + non_grain_columns
compare_df = compare_df[column_order]



compare_df['rate_plan_type_variance'] = compare_df.apply(lambda x: is_same_text(x['rate_plan_type_select'], x['rate_plan_type_table']), axis=1)
compare_df['replacement_rate_var %'] = compare_df.apply(lambda x: calc_variance_pct(x['replacement_rate_select'], x['replacement_rate_table']), axis=1)

append_df_to_existing_excel_workbook(compare_df, file_result, 'comparison')
sheet_format_dict['comparison'] = [{'#,##0': [8, 9], '0.0%': [11]}, 2, None]
# sheet_format_dict['comparison'] = [{'#,##0': [2, 3], '#,##0.00': [4, 5]}, 2, None]

#clean excel file
wb = openpyxl.load_workbook(filename=file_result)
format_excel_data(wb, sheet_format_dict, file_result)
clean_excel_file(wb, file_result)
wb.close()

time.sleep(4)

os.startfile(file_result)

print('process completed')

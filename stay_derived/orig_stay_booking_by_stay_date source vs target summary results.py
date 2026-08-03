from pathlib import Path
import pandas as pd
import os
import openpyxl
from openpyxl.utils import get_column_letter
import sys
from collections import Counter

project_dir = r'C:\Users\jwalker221\OneDrive - Hilton\Jira\Stay Derived\orig_stay_booking_by_stay_date'
search_pattern = r'orig_stay_booking_by_stay_date source and target table validation (* stay_id *).xlsx'


def create_excel_workbook(file_name_in):
    wb = openpyxl.Workbook()
    wb.save(file_name_in)
    wb.close()


def delete_file_if_exists(file_path_p):
    if os.path.exists(file_path_p):
        os.remove(file_path_p)


def append_df_to_existing_excel_workbook(df_p, workbook_path_p, sheet_name_p):
    with pd.ExcelWriter(workbook_path_p, mode='a', engine='openpyxl') as writer:
        df_p.to_excel(writer, sheet_name=sheet_name_p, index=False)


def clean_excel_file(wb_in, file_excel_p):
    for sheet in wb_in.sheetnames:
        if str(sheet)[0:5] == 'Sheet':
            wb_in.remove(wb_in[sheet])

    for sheet in wb_in.sheetnames:
        ws = wb_in[sheet]
        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max([len(str(cell.value) or "") for cell in column_cells])
            max_length = max_length + 3
            ws.column_dimensions[column_letter].width = max_length
            ws.freeze_panes = 'A2'
    wb_in.save(file_excel_p)


file_result = os.path.join(project_dir, f'source and target table validation summary results.xlsx')

delete_file_if_exists(file_result)
create_excel_workbook(file_result)

folder_path = Path(project_dir)
excel_files = list(folder_path.glob(search_pattern))

data = []
for file in excel_files:
    df_iter = pd.read_excel(file, sheet_name='variance')
    test_count = len(df_iter)
    variances = df_iter['is_variance'].to_numpy().tolist()
    counts = Counter(variances)
    variance_count = counts['yes']
    prop_and_stay = file.name.split('(')[1].split(')')[0]
    prop_cd = prop_and_stay.split(' stay_id ')[0]
    stay_id = prop_and_stay.split(' stay_id ')[1]
    data.append([file.name, prop_cd, stay_id, test_count, variance_count])

cols = ['file_name', 'prop_cd', 'stay_id', 'test_count', 'variance_count']
df_summary = pd.DataFrame(data, columns=cols)

append_df_to_existing_excel_workbook(df_summary, file_result, 'summary')

#clean excel file
wb = openpyxl.load_workbook(filename=file_result)
clean_excel_file(wb, file_result)
wb.close()

os.startfile(file_result)
print('success')

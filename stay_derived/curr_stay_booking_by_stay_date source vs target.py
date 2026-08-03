import pandas as pd
import os
import openpyxl
from openpyxl.utils import get_column_letter
import configparser
import redshift_connector as rc
import sys
import numpy as np

stay_id = '3079586382'
project_dir = r'C:\Users\jwalker221\OneDrive - Hilton\Jira\Stay Derived\curr_stay_booking_by_stay_date'
file_config = r'C:\Users\jwalker221\OneDrive - Hilton\Documents\cred.ini'
target_table = 'curr_stay_booking_by_stay_date'
file_sql_validation_name = 'curr_stay_booking_by_stay_date S2T validation (exclude rev columns) (stacked results) (dynamic).sql'
file_temp_name = 'temp.xlsx'

file_temp = os.path.join(project_dir, file_temp_name)
file_sql_validation = os.path.join(project_dir, file_sql_validation_name)

def connect_rs(host_p, database_p, user_p, password_p):
    conn = rc.connect(
        host=host_p
        , database=database_p
        , user=user_p
        , password=password_p
    )
    conn.autocommit = True
    return conn


def write_df_to_excel_temp_file(df, file_name, sheet_name='temp'):
    if os.path.exists(file_name):
        os.remove(file_name)
    df.to_excel(file_name, index=False, sheet_name=sheet_name)


def is_variance(row, columns_to_check, accepted_values):
    if row[columns_to_check].isin(accepted_values).all():
        return 'no'
    else:
        return 'yes'


def create_excel_workbook(file_name_in):
    wb = openpyxl.Workbook()
    wb.save(file_name_in)
    wb.close()


def delete_file_if_exists(file_path_p):
    if os.path.exists(file_path_p):
        os.remove(file_path_p)


def append_df_to_existing_excel_workbook(df_p, workbook_path_p, sheet_name_p):
    with pd.ExcelWriter(workbook_path_p, mode='a', engine='openpyxl') as writer:
        df_p.to_excel(writer, sheet_name=sheet_name_p, index=False)


def clean_excel_file(wb_in, file_excel_p):
    for sheet in wb_in.sheetnames:
        if str(sheet)[0:5] == 'Sheet':
            wb_in.remove(wb_in[sheet])

    for sheet in wb_in.sheetnames:
        ws = wb_in[sheet]
        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max([len(str(cell.value) or "") for cell in column_cells])
            max_length = max_length + 3
            ws.column_dimensions[column_letter].width = max_length
            ws.freeze_panes = 'A2'
    wb_in.save(file_excel_p)





config = configparser.ConfigParser()
config.read(file_config)
dw_config = config['dw_preprod']
host = dw_config['host']
dbname = dw_config['dbname']
user = dw_config['user']
password = dw_config['password']

rs_cred = [
    host
    ,dbname
    ,user
    ,password
]

rs_host, rs_database, rs_user, rs_pword = rs_cred
rs_conn = connect_rs(rs_host, rs_database, rs_user, rs_pword)
cursor = rs_conn.cursor()

with open(file_sql_validation, 'r') as file_r:
    sql = file_r.read()


sql = sql.replace('stay_id_variable', stay_id)
cursor.execute(sql)
df_detail = cursor.fetch_dataframe()
prop_cd = df_detail['prop_cd'].iloc[0]

file_result = os.path.join(project_dir, f'{target_table} source and target table validation ({prop_cd} stay_id {stay_id}).xlsx')

df_variance = df_detail[df_detail['row_source'] == 'VARIANCE'].copy()

df_column = df_variance.columns.tolist()
non_variance_column = ['row_source', 'stay_id', 'stay_dt', 'curr_booking_extract_last_update_dtm']
variance_column = [item for item in df_column if item not in set(non_variance_column)]
search_list = ['0', 'MATCH']
condition = df_variance[variance_column].isin(search_list).all(axis=1)
df_variance['is_variance'] = np.where(condition, 'no', 'yes')
column_order = ['is_variance'] + df_column
df_variance = df_variance[column_order]

delete_file_if_exists(file_result)
create_excel_workbook(file_result)

append_df_to_existing_excel_workbook(df_detail, file_result, 'detail')
append_df_to_existing_excel_workbook(df_variance, file_result, 'variance')


#clean excel file
wb = openpyxl.load_workbook(filename=file_result)
clean_excel_file(wb, file_result)
wb.close()

os.startfile(file_result)
print('success')
